import os
from openpyxl import load_workbook
from re import findall


def get_subject_from_address(address):
    for city in cities:
        if address.upper().find(city) != -1:
            return cities[city]


os.chdir("C:/Users/Dmitriy/Desktop/test_task")
file = 'Тестовое задание на позицию стажера-аналитика.xlsx'
file2 = 'subjects.xlsx'
wb = load_workbook(file)
wb2 = load_workbook(file2)
ws = wb.active
ws2 = wb2.active
sheet = wb.get_sheet_by_name('ja19 (1)')
sheet2 = wb2.get_sheet_by_name('1')
row_quantity = sheet.max_row
row_quantity2 = sheet2.max_row
indexes = dict()
cities = dict()

for i in range(2, row_quantity2 + 1):
    IND = ws2['A' + str(i)].value
    RF_SUBJECT = ws2['B' + str(i)].value.upper().strip()
    CITY = ws2['C' + str(i)].value
    indexes[IND] = RF_SUBJECT
    cities[CITY] = RF_SUBJECT

for i in range(2, row_quantity + 1):
    cell_value =  ws['C' + str(i)].value
    if cell_value[-3:].replace(',', '').strip().lower() == 'ru':
        ws['E' + str(i)] = ''.join(findall('\d', cell_value[-10:-4]))[0:3]
        ws['F' + str(i)] = indexes.get(sheet.cell(row=i, column=5).value)
        if ws['F' + str(i)].value is None:
            adds = ws['C' + str(i)].value
            ws['F' + str(i)] = get_subject_from_address(adds)
wb.save(file)
